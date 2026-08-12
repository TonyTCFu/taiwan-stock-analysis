import os
import json
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

json_path = "/Users/TonyFu/.gemini/antigravity/scratch/taiwan-stock-dashboard/data/stock_data.json"
with open(json_path, encoding="utf-8") as f:
    stock_payload = json.load(f)

updated_at = stock_payload.get("updated_at", "")
data_source = stock_payload.get("data_source", "永丰金 Shioaji API")
stocks = stock_payload.get("stocks", [])

desktop_dir = "/Users/TonyFu/Desktop/台湾股市分析"
os.makedirs(desktop_dir, exist_ok=True)
excel_filename = "台股 5 大 AI 核心卡位龙头深度投资与买卖点筹码量化分析表.xlsx"
target_excel_path = os.path.join(desktop_dir, excel_filename)
artifact_excel_path = "/Users/TonyFu/.gemini/antigravity/brain/236b9bdf-d399-4da3-8f7f-3f6b26ffcc9f/台股核心卡位龙头5强深度投资与买卖点分析表.xlsx"

wb = openpyxl.Workbook()

font_title = Font(name="Microsoft YaHei", size=16, bold=True, color="FFFFFF")
font_subtitle = Font(name="Microsoft YaHei", size=10, italic=True, color="333333")
font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
font_cell_zh = Font(name="Microsoft YaHei", size=10, color="262626")
font_cell_code = Font(name="Calibri", size=11, bold=True, color="1F4E79")
font_price_ref = Font(name="Calibri", size=11, bold=True, color="C00000")
font_buy_zone = Font(name="Microsoft YaHei", size=10, bold=True, color="1B365D")
font_sell_zone = Font(name="Microsoft YaHei", size=10, bold=True, color="2E75B6")
font_target_price = Font(name="Calibri", size=11, bold=True, color="385723")
font_stop_loss = Font(name="Calibri", size=11, bold=True, color="C00000")

fill_title = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
fill_subtitle = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
fill_header = PatternFill(start_color="2C4D75", end_color="2C4D75", fill_type="solid")
fill_zebra = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

thin_side = Side(style='thin', color='D9D9D9')
border_grid = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

ws1 = wb.active
ws1.title = "核心卡位龙头全景深度分析"
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells('A1:O1')
t1 = ws1['A1']
t1.value = "台股 5 大 AI 核心卡位龙头深度投资与买卖点筹码量化分析表"
t1.font = font_title
t1.alignment = Alignment(horizontal="center", vertical="center")
t1.fill = fill_title
ws1.row_dimensions[1].height = 42

ws1.merge_cells('A2:O2')
s1 = ws1['A2']
s1.value = f"定盘时间: {updated_at} | 数据源: {data_source} | 包含三大法人买卖超、财报发布日及主力资金分析"
s1.font = font_subtitle
s1.alignment = Alignment(horizontal="center", vertical="center")
s1.fill = fill_subtitle
ws1.row_dimensions[2].height = 24

headers1 = [
    "股票代码", "公司名称", "核心卡位定位", "当日开/高/低/收盘 (Shioaji)",
    "获利率指标 (毛利/净利/ROE/EPS)", "财报公布日/预估", "三大法人买卖超 (外资/投信)",
    "主力资金进出与筹码沉淀", "选择理由 (Serenity 护城河)", "参考最新价",
    "建议买入区间", "阶段止盈/卖出区间", "长线目标价", "防守止损价", "单日成交量"
]

ws1.row_dimensions[3].height = 36
for col_idx, h in enumerate(headers1, 1):
    c = ws1.cell(row=3, column=col_idx, value=h)
    c.fill = fill_header
    c.font = font_header
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx, s in enumerate(stocks, 4):
    ws1.row_dimensions[row_idx].height = 105
    is_zebra = (row_idx % 2 == 1)
    
    ohlc_str = f"开盘: {s['open_price']}\n最高: {s['high_price']}\n最低: {s['low_price']}\n收盘: {s['last_price']}"
    profit_str = f"• 毛利率：{s['gross_margin']}\n• 净利率：{s['net_margin']}\n• ROE：{s['roe']}\n• EPS：{s['eps_single']}"
    inst_str = f"外资: {s['institutional_flow']['foreign_net']}\n投信: {s['institutional_flow']['trust_net']}\n{s['institutional_flow']['summary']}"
    buy_str = f"建议买点：{s['buy_zone_sub']}\n黄金重仓：{s['buy_zone_heavy']}"
    
    row_data = [
        s['symbol'],
        s['name'],
        s['chokepoint'],
        ohlc_str,
        profit_str,
        s['earnings_date'],
        inst_str,
        f"{s['capital_inflow']['large_order_ratio']}\n{s['capital_inflow']['capital_status']}",
        f"【Chokepoint {s['score']}分】卡位最上游铲子卖家，客户切换周期超 12-36 个月，垄断护城河极深。",
        f"NT$ {s['last_price']}",
        buy_str,
        s['take_profit'],
        s['target_price'],
        s['stop_loss'],
        f"{s['volume']:,} 张"
    ]
    
    for col_idx, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border_grid
        cell.font = font_cell_zh
        if is_zebra:
            cell.fill = fill_zebra
            
        if col_idx == 1:
            cell.font = font_cell_code
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [2, 4, 6]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx == 2:
                cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F4E79")
        elif col_idx == 10:
            cell.font = font_price_ref
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx in [11, 12, 13, 14, 15]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_idx == 11:
                cell.font = font_buy_zone
            elif col_idx == 12:
                cell.font = font_sell_zone
            elif col_idx == 13:
                cell.font = font_target_price
            elif col_idx == 14:
                cell.font = font_stop_loss
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

col1_widths = {1: 13, 2: 18, 3: 26, 4: 20, 5: 26, 6: 25, 7: 30, 8: 30, 9: 30, 10: 16, 11: 26, 12: 20, 13: 18, 14: 16, 15: 16}
for col_idx, width in col1_widths.items():
    ws1.column_dimensions[get_column_letter(col_idx)].width = width

# SHEET 2: 买卖决策速查面板
ws2 = wb.create_sheet(title="买卖决策速查面板")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells('A1:I1')
t2 = ws2['A1']
t2.value = "台股 5 大卡位龙头 Shioaji 实时买卖点与风控决策面板"
t2.font = font_title
t2.alignment = Alignment(horizontal="center", vertical="center")
t2.fill = fill_title
ws2.row_dimensions[1].height = 42

ws2.merge_cells('A2:I2')
s2 = ws2['A2']
s2.value = f"数据源: 永丰金 Shioaji 实时盘口 | 定盘时间: {updated_at} | 策略: 3-3-4 分批建仓"
s2.font = font_subtitle
s2.alignment = Alignment(horizontal="center", vertical="center")
s2.fill = fill_subtitle
ws2.row_dimensions[2].height = 24

headers2 = ["股票代码", "公司名称", "开盘价", "最高价", "最低价", "收盘(最新)价", "建议买入区间", "波段止盈位", "防守止损价"]
ws2.row_dimensions[3].height = 32
for col_idx, h in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=col_idx, value=h)
    c.fill = fill_header
    c.font = font_header
    c.alignment = Alignment(horizontal="center", vertical="center")

for row_idx, s in enumerate(stocks, 4):
    ws2.row_dimensions[row_idx].height = 36
    is_zebra = (row_idx % 2 == 1)
    
    row_data = [
        s['symbol'], s['name'], f"NT$ {s['open_price']}", f"NT$ {s['high_price']}",
        f"NT$ {s['low_price']}", f"NT$ {s['last_price']}", s['buy_zone_sub'], s['take_profit'], s['stop_loss']
    ]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border_grid
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if is_zebra:
            cell.fill = fill_zebra
            
        if col_idx == 1:
            cell.font = font_cell_code
        elif col_idx == 2:
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="1F4E79")
        elif col_idx == 6:
            cell.font = font_price_ref
        elif col_idx == 7:
            cell.font = font_buy_zone
        elif col_idx == 8:
            cell.font = font_sell_zone
        elif col_idx == 9:
            cell.font = font_stop_loss

col2_widths = {1: 14, 2: 18, 3: 16, 4: 16, 5: 16, 6: 18, 7: 24, 8: 20, 9: 16}
for col_idx, width in col2_widths.items():
    ws2.column_dimensions[get_column_letter(col_idx)].width = width

wb.save(target_excel_path)
shutil.copyfile(target_excel_path, artifact_excel_path)
print(f"Shioaji 财报与三大法人 Excel 更新完成: {target_excel_path}")
